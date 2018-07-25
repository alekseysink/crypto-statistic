import datetime
from dateutil import parser
from django.core.management.base import BaseCommand
from openpyxl import load_workbook, Workbook

from cryptocurrencies.apps.Currencies.models import CryptoCurrency, CryptoCurrencyRateData, \
    CryptoCurrencyRateDataChanges


class Command(BaseCommand):
    help = 'Base Command'

    def handle(self, *args, **options):

        # workbook = load_workbook('source.xlsx')
        # for sheet in workbook.worksheets:
        #     crypto_currency = CryptoCurrency.objects.get(symbol=sheet.title)
        #     for row in sheet.iter_rows(row_offset=2):
        #         try:
        #             CryptoCurrencyRateData.objects.create(
        #                 crypto_currency=crypto_currency,
        #                 date=row[0].value,
        #                 open_rate=float(row[1].value),
        #                 close_rate=float(row[4].value),
        #                 high_rate=float(row[2].value),
        #                 low_rate=float(row[3].value),
        #                 adj_close=float(row[5].value),
        #                 volume=float(row[6].value)
        #             )
        #         except Exception:
        #             pass
        #         print(
        #             f"Name: {sheet.title} "
        #             f"Date: {row[0].value} "
        #             f"Open: {row[1].value} "
        #             f"High: {row[2].value} "
        #             f"Low: {row[3].value} "
        #             f"Close: {row[4].value} "
        #             f"Adj Close {row[5].value} "
        #             f"Volume: {row[6].value}"
        #         )

        date = CryptoCurrencyRateDataChanges.objects.all().latest('date').date

        currency_list = list(CryptoCurrencyRateDataChanges.objects.filter(date=date).values_list('crypto_currency__symbol', flat=True))


        workbook = Workbook()
        worksheet = workbook.create_sheet()

        headers = ['Date'] + currency_list
        worksheet.append(headers)

        crypto_currency_data_changes = CryptoCurrencyRateDataChanges.objects.filter(date=date)

        while not (date.year == 2017 and date.day == 21 and date.month == 6):
            date = date - datetime.timedelta(days=1)
            crypto_currency_data_changes = CryptoCurrencyRateDataChanges.objects.filter(date=date)
            row_key_value = {
                'Date': date,
            }
            for crypto in crypto_currency_data_changes:
                row_key_value[crypto.crypto_currency.symbol] = crypto.changes
            crypto_data = [row_key_value.get(header, 0) for header in headers[1:]]
            worksheet.append([
                date,
            ] + crypto_data)
            print([date] + crypto_data)
        workbook.save('result.xlsx')
