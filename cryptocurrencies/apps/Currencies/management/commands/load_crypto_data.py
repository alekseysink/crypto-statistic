import datetime
from dateutil import parser
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from cryptocurrencies.apps.Currencies.models import CryptoCurrency, CryptoCurrencyRateData, \
    CryptoCurrencyRateDataChanges


class Command(BaseCommand):
    help = 'Base Command'

    def handle(self, *args, **options):

        # workbook = load_workbook('source.xlsx')
        # for sheet in workbook.worksheets:
        #     crypto_currency = CryptoCurrency.objects.get(symbol=sheet.title)
        #     for row in sheet.iter_rows(row_offset=2):
        #         try:
        #             CryptoCurrencyRateData.objects.create(
        #                 crypto_currency=crypto_currency,
        #                 date=row[0].value,
        #                 open_rate=float(row[1].value),
        #                 close_rate=float(row[4].value),
        #                 high_rate=float(row[2].value),
        #                 low_rate=float(row[3].value),
        #                 adj_close=float(row[5].value),
        #                 volume=float(row[6].value)
        #             )
        #         except Exception:
        #             pass
        #         print(
        #             f"Name: {sheet.title} "
        #             f"Date: {row[0].value} "
        #             f"Open: {row[1].value} "
        #             f"High: {row[2].value} "
        #             f"Low: {row[3].value} "
        #             f"Close: {row[4].value} "
        #             f"Adj Close {row[5].value} "
        #             f"Volume: {row[6].value}"
        #         )
        for crypto_currency in CryptoCurrency.objects.all():
            rates_data = CryptoCurrencyRateData.objects.filter(crypto_currency=crypto_currency).order_by('-date')
            for idx, rate_data in enumerate(rates_data):
                try:
                    CryptoCurrencyRateDataChanges.objects.create(
                        date=rate_data.date,
                        crypto_currency=crypto_currency,
                        changes=float((rate_data.close_rate - rates_data[idx+1].close_rate)/rates_data[idx+1].close_rate)
                    )
                except Exception as e:
                    print(e)